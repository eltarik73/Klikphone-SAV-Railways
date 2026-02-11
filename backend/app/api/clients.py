"""
API CRUD Clients — compatible schéma PostgreSQL existant.
"""

import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from app.database import get_cursor
from app.models import ClientCreate, ClientUpdate, ClientOut
from app.api.auth import get_current_user

router = APIRouter(prefix="/api/clients", tags=["clients"])


@router.get("", response_model=list[ClientOut])
async def list_clients(
    search: Optional[str] = None,
    limit: int = Query(100, le=500),
    offset: int = 0,
    user: dict = Depends(get_current_user),
):
    """Liste les clients avec recherche optionnelle."""
    if search:
        with get_cursor() as cur:
            s = f"%{search}%"
            cur.execute("""
                SELECT * FROM clients
                WHERE nom ILIKE %s OR prenom ILIKE %s
                      OR telephone LIKE %s OR email ILIKE %s
                      OR societe ILIKE %s
                ORDER BY date_creation DESC
                LIMIT %s OFFSET %s
            """, (s, s, s, s, s, limit, offset))
            return cur.fetchall()
    else:
        with get_cursor() as cur:
            cur.execute(
                "SELECT * FROM clients ORDER BY date_creation DESC LIMIT %s OFFSET %s",
                (limit, offset),
            )
            return cur.fetchall()


# ─── EXPORT routes (MUST be before /{client_id}) ──────────────
@router.get("/export/csv")
async def export_clients_csv(user: dict = Depends(get_current_user)):
    """Exporte tous les clients au format CSV avec nb de tickets."""
    with get_cursor() as cur:
        cur.execute("""
            SELECT c.id, c.nom, c.prenom, c.telephone, c.email, c.societe,
                   c.carte_camby, c.date_creation,
                   COUNT(t.id) AS nb_tickets
            FROM clients c
            LEFT JOIN tickets t ON t.client_id = c.id
            GROUP BY c.id
            ORDER BY c.nom
        """)
        rows = cur.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nom", "Prénom", "Téléphone", "Email", "Société", "Carte Camby", "Date création", "Nb tickets"])
    for r in rows:
        writer.writerow([r["id"], r["nom"], r["prenom"], r["telephone"], r["email"], r["societe"], r["carte_camby"], r["date_creation"], r["nb_tickets"]])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clients_klikphone.csv"},
    )


@router.get("/export/excel")
async def export_clients_excel(user: dict = Depends(get_current_user)):
    """Exporte tous les clients au format Excel (.xlsx) avec nb de tickets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_cursor() as cur:
        cur.execute("""
            SELECT c.id, c.nom, c.prenom, c.telephone, c.email, c.societe,
                   c.carte_camby, c.date_creation,
                   COUNT(t.id) AS nb_tickets
            FROM clients c
            LEFT JOIN tickets t ON t.client_id = c.id
            GROUP BY c.id
            ORDER BY c.nom
        """)
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = ["ID", "Nom", "Prénom", "Téléphone", "Email", "Société", "Carte Camby", "Date création", "Nb tickets"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, r in enumerate(rows, 2):
        vals = [r["id"], r["nom"], r["prenom"], r["telephone"], r["email"], r["societe"], r["carte_camby"], str(r["date_creation"] or ""), r["nb_tickets"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clients_klikphone.xlsx"},
    )


@router.get("/{client_id}", response_model=ClientOut)
async def get_client(client_id: int, user: dict = Depends(get_current_user)):
    """Récupère un client par ID."""
    with get_cursor() as cur:
        cur.execute("SELECT * FROM clients WHERE id = %s", (client_id,))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "Client non trouvé")
    return row


@router.get("/tel/{telephone}")
async def get_client_by_tel(telephone: str):
    """Recherche un client par téléphone (public — formulaire client)."""
    with get_cursor() as cur:
        cur.execute("SELECT * FROM clients WHERE telephone = %s", (telephone,))
        row = cur.fetchone()
    return row  # None si pas trouvé (pas d'erreur 404)


@router.post("", response_model=ClientOut)
async def create_or_get_client(data: ClientCreate):
    """Crée un client ou retourne l'existant si le téléphone existe déjà.
    
    Compatible avec get_or_create_client de l'app Streamlit.
    """
    with get_cursor() as cur:
        # Vérifier si le client existe
        cur.execute("SELECT * FROM clients WHERE telephone = %s", (data.telephone,))
        existing = cur.fetchone()

        if existing:
            # Mettre à jour les infos si elles ont changé
            cur.execute("""
                UPDATE clients SET 
                    nom = COALESCE(NULLIF(%s, ''), nom),
                    prenom = COALESCE(NULLIF(%s, ''), prenom),
                    email = COALESCE(NULLIF(%s, ''), email),
                    societe = COALESCE(NULLIF(%s, ''), societe),
                    carte_camby = %s
                WHERE id = %s
                RETURNING *
            """, (
                data.nom, data.prenom, data.email,
                data.societe, data.carte_camby, existing["id"],
            ))
            return cur.fetchone()

        # Créer le client
        cur.execute("""
            INSERT INTO clients (nom, prenom, telephone, email, societe, carte_camby)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING *
        """, (
            data.nom, data.prenom, data.telephone,
            data.email, data.societe, data.carte_camby,
        ))
        return cur.fetchone()


@router.patch("/{client_id}", response_model=dict)
async def update_client(
    client_id: int,
    data: ClientUpdate,
    user: dict = Depends(get_current_user),
):
    """Met à jour un client."""
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        return {"ok": True}

    set_clause = ", ".join(f"{k} = %s" for k in updates.keys())
    values = list(updates.values()) + [client_id]

    with get_cursor() as cur:
        cur.execute(f"UPDATE clients SET {set_clause} WHERE id = %s", values)

    return {"ok": True}


@router.delete("/{client_id}", response_model=dict)
async def delete_client(client_id: int, user: dict = Depends(get_current_user)):
    """Supprime un client (vérifie qu'il n'a pas de tickets)."""
    with get_cursor() as cur:
        cur.execute("SELECT COUNT(*) as cnt FROM tickets WHERE client_id = %s", (client_id,))
        row = cur.fetchone()
        if row and row["cnt"] > 0:
            raise HTTPException(400, f"Ce client a {row['cnt']} ticket(s). Supprimez-les d'abord.")

        cur.execute("DELETE FROM clients WHERE id = %s", (client_id,))

    return {"ok": True}


@router.get("/{client_id}/tickets", response_model=list)
async def get_client_tickets(client_id: int, user: dict = Depends(get_current_user)):
    """Récupère tous les tickets d'un client."""
    with get_cursor() as cur:
        cur.execute(
            "SELECT * FROM tickets WHERE client_id = %s ORDER BY date_depot DESC",
            (client_id,),
        )
        return cur.fetchall()


